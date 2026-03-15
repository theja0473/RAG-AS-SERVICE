"""Data ingestion agent for extracting text from various sources.

This module provides functions for extracting text from PDF, DOCX,
XLSX, TXT files, and URLs.
"""

from typing import Dict, Any
from pathlib import Path

import pdfplumber
from docx import Document
from openpyxl import load_workbook
import httpx
from bs4 import BeautifulSoup


def extract_pdf_text(file_path: str) -> str:
    """Extract text from PDF file.

    Args:
        file_path: Path to PDF file.

    Returns:
        Extracted text content.
    """
    text_parts = []

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)

    return "\n\n".join(text_parts)


def extract_docx_text(file_path: str) -> str:
    """Extract text from DOCX file.

    Args:
        file_path: Path to DOCX file.

    Returns:
        Extracted text content.
    """
    doc = Document(file_path)
    text_parts = []

    for paragraph in doc.paragraphs:
        if paragraph.text.strip():
            text_parts.append(paragraph.text)

    return "\n\n".join(text_parts)


def extract_xlsx_text(file_path: str) -> str:
    """Extract text from XLSX file.

    Args:
        file_path: Path to XLSX file.

    Returns:
        Extracted text content.
    """
    workbook = load_workbook(file_path, read_only=True)
    text_parts = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text_parts.append(f"Sheet: {sheet_name}")

        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
            if row_text.strip():
                text_parts.append(row_text)

    workbook.close()
    return "\n".join(text_parts)


def extract_txt_text(file_path: str) -> str:
    """Extract text from TXT file.

    Args:
        file_path: Path to TXT file.

    Returns:
        Extracted text content.
    """
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def extract_url_text(url: str) -> str:
    """Extract text from URL.

    Args:
        url: URL to fetch and extract text from.

    Returns:
        Extracted text content.
    """
    response = httpx.get(url, timeout=30.0, follow_redirects=True)
    response.raise_for_status()

    soup = BeautifulSoup(response.content, "html.parser")

    # Remove script and style elements
    for script in soup(["script", "style"]):
        script.decompose()

    # Get text
    text = soup.get_text()

    # Clean up text
    lines = (line.strip() for line in text.splitlines())
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    text = "\n".join(chunk for chunk in chunks if chunk)

    return text


def extract_text_from_file(file_path: str, source_type: str) -> Dict[str, Any]:
    """Extract text from file based on type.

    Args:
        file_path: Path to file.
        source_type: Type of file (pdf, docx, xlsx, txt).

    Returns:
        Dictionary with extracted text and metadata.
    """
    extractors = {
        "pdf": extract_pdf_text,
        "docx": extract_docx_text,
        "xlsx": extract_xlsx_text,
        "txt": extract_txt_text,
    }

    if source_type not in extractors:
        raise ValueError(f"Unsupported source type: {source_type}")

    text = extractors[source_type](file_path)

    return {
        "text": text,
        "metadata": {
            "source": Path(file_path).name,
            "source_type": source_type,
            "file_path": file_path,
        }
    }


def extract_text_from_url(url: str) -> Dict[str, Any]:
    """Extract text from URL.

    Args:
        url: URL to extract from.

    Returns:
        Dictionary with extracted text and metadata.
    """
    text = extract_url_text(url)

    return {
        "text": text,
        "metadata": {
            "source": url,
            "source_type": "url",
        }
    }
